import openpyxl
from django.http import HttpResponse
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.contrib import messages
from inventory.forms import CategoryForm, DepartementForm, EmployeeForm, ItemForm
from inventory.models import Category, Departement, Employee, Item, ItemUnit
from openpyxl.styles import Font, Alignment


##########################################################  views for category
class CategoryListView(ListView):
    model = Category
    template_name = "inventory/category/list.html"

    def get_context_data(self, **kwrags):
        context = super().get_context_data(**kwrags)
        context['title']="List Category"
        return context


class CategoryCreateView(CreateView):
    model = Category
    form_class = CategoryForm
    template_name = "inventory/category/create.html"
    success_url = reverse_lazy('inventory:list_category')

    def get_context_data(self, **kwrags):
        context = super().get_context_data(**kwrags)
        context['title']="Create Category"
        return context


class CategoryUpdateView(UpdateView):
    model = Category
    form_class = CategoryForm
    template_name = "inventory/category/create.html"
    success_url = reverse_lazy('inventory:list_category')

    def get_context_data(self, **kwrags):
        context = super().get_context_data(**kwrags)
        context['title']="Update Category"
        return context


class CategoryDeleteView(DeleteView):
    model = Category
    template_name = "inventory/category/delete.html"
    success_url = reverse_lazy('inventory:list_category')

    def delete(self, request, *args, **kwargs):
        category = self.get_object()
        messages.success(self.request, f"Category '{category.name}' berhasil disimpan.")
        return super().delete(request, *args, **kwargs)
                                                    
def categoryToExcel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "List Category"

    # Judul besar di baris 1
    ws.merge_cells('A1:B1')  # gabungkan dari kolom A sampai B
    ws['A1'] = "List Category"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Header
    headers = ['Id', 'Name']
    ws.append(headers)

    # Data
    for idx, s in enumerate(Category.objects.all(), start=1):
        ws.append([
            idx,
            s.name,
        ])

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=category.xlsx'
    wb.save(response)
    return response

##########################################################  views for departemen
class DepartementListView(ListView):
    model = Departement
    template_name = "inventory/departement/list.html"

    def get_context_data(self, **kwrags):
        context = super().get_context_data(**kwrags)
        context['title']="List Departement"
        return context


class DepartementCreateView(CreateView):
    model = Departement
    form_class = DepartementForm
    template_name = "inventory/departement/create.html"
    success_url = reverse_lazy('inventory:list_departement')

    def get_context_data(self, **kwrags):
        context = super().get_context_data(**kwrags)
        context['title']="Create Departement"
        return context


class DepartementUpdateView(UpdateView):
    model = Departement
    form_class = DepartementForm
    template_name = "inventory/departement/create.html"
    success_url = reverse_lazy('inventory:list_departement')

    def get_context_data(self, **kwrags):
        context = super().get_context_data(**kwrags)
        context['title']="Update Departement"
        return context


class DepartementDeleteView(DeleteView):
    model = Departement
    template_name = "inventory/departement/delete.html"
    success_url = reverse_lazy('inventory:list_departement')

    def delete(self, request, *args, **kwargs):
        departement = self.get_object()
        messages.success(self.request, f"departement '{departement.name}' berhasil disimpan.")
        return super().delete(request, *args, **kwargs)
                                                    
def departementToExcel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "List Category"

    # Judul besar di baris 1
    ws.merge_cells('A1:C1')  # gabungkan dari kolom A sampai B
    ws['A1'] = "List Departement"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Header
    headers = ['Id', 'Name', 'Leader']
    ws.append(headers)

    # Data
    for idx, s in enumerate(Departement.objects.all(), start=1):
        ws.append([
            idx,
            s.name,
            s.leader,
        ])

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=departement.xlsx'
    wb.save(response)
    return response

##########################################################  views for employee
class EmployeeListView(ListView):
    model = Employee
    template_name = "inventory/employee/list.html"

    def get_context_data(self, **kwrags):
        context = super().get_context_data(**kwrags)
        context['title']="List Employee"
        return context


class EmployeeCreateView(CreateView):
    model = Employee
    form_class = EmployeeForm
    template_name = "inventory/employee/create.html"
    success_url = reverse_lazy('inventory:list_employee')

    def get_context_data(self, **kwrags):
        context = super().get_context_data(**kwrags)
        context['title']="Create Employee"
        return context


class EmployeeUpdateView(UpdateView):
    model = Employee
    form_class = EmployeeForm
    template_name = "inventory/employee/create.html"
    success_url = reverse_lazy('inventory:list_employee')

    def get_context_data(self, **kwrags):
        context = super().get_context_data(**kwrags)
        context['title']="Update Employee"
        return context


class EmployeeDeleteView(DeleteView):
    model = Employee
    template_name = "inventory/employee/delete.html"
    success_url = reverse_lazy('inventory:list_employee')

    def delete(self, request, *args, **kwargs):
        Employee = self.get_object()
        messages.success(self.request, f"Employee '{Employee.name}' berhasil disimpan.")
        return super().delete(request, *args, **kwargs)
                                                    
def employeeToExcel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "List Employee"

    # Judul besar di baris 1
    ws.merge_cells('A1:G1')  # gabungkan dari kolom A sampai B
    ws['A1'] = "List Employee"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Header
    headers = ['Id', 'User', 'Employee Number', 'Name', 'Departement', 'Position', 'email', 'phone']
    ws.append(headers)

    # Data
    for idx, s in enumerate(Employee.objects.all(), start=1):
        ws.append([
            idx,
            s.user,
            s.employee_id,
            s.name,
            s.departement.name,
            s.position,
            s.email,
            s.phone,
        ])

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Employee.xlsx'
    wb.save(response)
    return response

##########################################################  views for item
class ItemListView(ListView):
    model = Item
    template_name = "inventory/item/list.html"

    def get_context_data(self, **kwrags):
        context = super().get_context_data(**kwrags)
        context['title']="List Item"
        return context


class ItemCreateView(CreateView):
    model = Item
    form_class = ItemForm
    template_name = "inventory/item/create.html"
    success_url = reverse_lazy('inventory:list_employee')

    def get_context_data(self, **kwrags):
        context = super().get_context_data(**kwrags)
        context['title']="Create Employee"
        return context


class ItemUpdateView(UpdateView):
    model = Item
    form_class = ItemForm
    template_name = "inventory/item/create.html"
    success_url = reverse_lazy('inventory:list_item')

    def get_context_data(self, **kwrags):
        context = super().get_context_data(**kwrags)
        context['title']="Update Item"
        return context


class ItemDeleteView(DeleteView):
    model = Item
    template_name = "inventory/item/delete.html"
    success_url = reverse_lazy('inventory:list_item')

    def delete(self, request, *args, **kwargs):
        Employee = self.get_object()
        messages.success(self.request, f"Item '{Employee.name}' berhasil dihapus.")
        return super().delete(request, *args, **kwargs)
                                                    
def itemToExcel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "List Item"

    # Judul besar di baris 1
    ws.merge_cells('A1:G1')  # gabungkan dari kolom A sampai B
    ws['A1'] = "List Item"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Header
    headers = ['Id', 'SKU', 'Category', 'Name', 'Brand', 'Model', 'Cpu', 'Ram', 'Storage', 'Display', 'Os','Entry Date']
    ws.append(headers)

    # Data
    for idx, s in enumerate(Employee.objects.all(), start=1):
        ws.append([
            idx,
            s.sku,
            s.category,
            s.name,
            s.brand,
            s.model,
            s.cpu,
            s.ram,
            s.storage,
            s.display,
            s.os,
            s.entry_date,
        ])

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Item.xlsx'
    wb.save(response)
    return response

##########################################################  views for item
class ItemUnitListView(ListView):
    model = ItemUnit
    template_name = "inventory/item_unit/list.html"

    def get_context_data(self, **kwrags):
        context = super().get_context_data(**kwrags)
        context['title']="List Item Unit"
        return context


class ItemUnitCreateView(CreateView):
    model = Item
    form_class = ItemForm
    template_name = "inventory/item_unit/create.html"
    success_url = reverse_lazy('inventory:list_employee')

    def get_context_data(self, **kwrags):
        context = super().get_context_data(**kwrags)
        context['title']="Create Employee"
        return context


class ItemUpdateView(UpdateView):
    model = Item
    form_class = ItemForm
    template_name = "inventory/item_unit/create.html"
    success_url = reverse_lazy('inventory:list_item')

    def get_context_data(self, **kwrags):
        context = super().get_context_data(**kwrags)
        context['title']="Update Item"
        return context


class ItemDeleteView(DeleteView):
    model = Item
    template_name = "inventory/item_unit/delete.html"
    success_url = reverse_lazy('inventory:list_item')

    def delete(self, request, *args, **kwargs):
        Employee = self.get_object()
        messages.success(self.request, f"Item '{Employee.name}' berhasil dihapus.")
        return super().delete(request, *args, **kwargs)
                                                    
def itemToExcel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "List Item"

    # Judul besar di baris 1
    ws.merge_cells('A1:G1')  # gabungkan dari kolom A sampai B
    ws['A1'] = "List Item"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Header
    headers = ['Id', 'SKU', 'Category', 'Name', 'Brand', 'Model', 'Cpu', 'Ram', 'Storage', 'Display', 'Os','Entry Date']
    ws.append(headers)

    # Data
    for idx, s in enumerate(Employee.objects.all(), start=1):
        ws.append([
            idx,
            s.sku,
            s.category,
            s.name,
            s.brand,
            s.model,
            s.cpu,
            s.ram,
            s.storage,
            s.display,
            s.os,
            s.entry_date,
        ])

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Item.xlsx'
    wb.save(response)
    return response